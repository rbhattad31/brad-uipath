import pandas as pd
from datetime import datetime
import numpy
import openpyxl
from openpyxl.styles import Font, PatternFill

print("starting at " + datetime.now().strftime("%H:%M:%S"))

excel_path = "purchase_registers_raw.xlsx"
previous_quarter_file = "Previous_Quarter_Final_File.xlsx"
excel_file_as_pd = pd.read_excel(excel_path, sheet_name="Sheet1")


def generate_purchase_type_wise(excel_file_pd, previous_quarter_final_file):
    # create pivot table
    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")

    # reset "indices created during pivot table creation" - for merging
    purchase_type_wise_pd = purchase_type_wise_pd.reset_index()

    # read previous quarters final working file - pd will be replaced with Nan in any blank cells
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D")

    # replace Nan with blank
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)

    # merging present and previous quarter purchase type wise data -  pd will be replaced with Nan in any blank cells
    purchase_type_wise_comparatives_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])

    # replacing all Nan's with zeros in Present and previous Quarter's values columns
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    columns_names = purchase_type_wise_comparatives_pd.columns.values.tolist()
    # returns as ['Valuation Class', 'Valuation Class Text', 'GR Amt.in loc.cur.', 'Previous Quarter']

    # dropping columns present and previous quarters both have values as zero
    purchase_type_wise_comparatives_pd.drop(purchase_type_wise_comparatives_pd.index[(purchase_type_wise_comparatives_pd[columns_names[2]] == 0) & (purchase_type_wise_comparatives_pd[columns_names[3]] == 0)], inplace=True)

    # create a new column - Success
    purchase_type_wise_comparatives_pd['Variance'] = 0

    pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

    # variance formula implementation using index
    for index in purchase_type_wise_comparatives_pd.index:
        present_quarter = purchase_type_wise_comparatives_pd[columns_names[2]][index]
        previous_quarter = purchase_type_wise_comparatives_pd[columns_names[3]][index]
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        purchase_type_wise_comparatives_pd['Variance'][index] = variance

    # sorting
    # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
    grand_total = purchase_type_wise_comparatives_pd[columns_names[2]].values[-1]
    purchase_type_wise_comparatives_pd[columns_names[2]].values[-1] = 0
    purchase_type_wise_comparatives_pd.sort_values(by=columns_names[2], axis=0, ascending=False, inplace=True)

    purchase_type_wise_comparatives_pd[columns_names[2]].values[-1] = grand_total

    return purchase_type_wise_comparatives_pd


def generate_month_wise(excel_file_pd, previous_quarter_final_file):

    #  Create Month column
    excel_file_pd['Month'] = pd.DatetimeIndex(excel_file_pd['GR Posting Date']).month_name().str[:3]

    #  selecting only required columns
    excel_file_pd = excel_file_pd[["GR Posting Date", "GR Amt.in loc.cur.", "Month"]]

    # create pivot table, sort = False to not sort Month column as per alphabetical order - mandatory
    month_wise_pd = pd.pivot_table(excel_file_pd, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=False)

    # reset month index after pivot table creation for concatenation
    month_wise_pd = month_wise_pd.reset_index()

    # read from previous quarters final working file
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Month Wise Comparatives", usecols="C,D")

    # concatenation instead of merge as there are no common Columns to merge.
    month_wise_comparatives_pd = pd.concat([month_wise_pd, previous_quarter_final_file_pd], axis=1)

    columns_names = month_wise_comparatives_pd.columns.values.tolist()
    # returns as ['Month', 'GR Amt.in loc.cur.', 'Month.1', 'Previous Quarter as Q3 FY 21-22']

    # create a new column - Success
    month_wise_comparatives_pd['Variance'] = 0

    pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

    # variance formula implementation using index
    for index in month_wise_comparatives_pd.index:
        present_quarter = month_wise_comparatives_pd[columns_names[1]][index]
        previous_quarter = month_wise_comparatives_pd[columns_names[3]][index]
        variance = (present_quarter - previous_quarter) / previous_quarter
        month_wise_comparatives_pd['Variance'][index] = variance

    return month_wise_comparatives_pd


def generate_plant_type_wise(excel_file_pd, previous_quarter_final_file):

    # create pivot table
    plant_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Plant"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=True)

    # reset index created after pivot table creation for merging
    plant_type_wise_pd = plant_type_wise_pd.reset_index()

    # read previous quarters final working file
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Plant Wise Comparatives", usecols="A,C")

    # merging present and previous quarter purchase type wise data
    plant_type_wise_comparatives_pd = pd.merge(plant_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Plant"])

    # replacing all Nan's with zeros in Present and previous values columns - not necessary now but if a new plant is added
    # Nan's will be formed in previous quarter columns, eliminates NaN error
    plant_type_wise_comparatives_pd = plant_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    columns_names = plant_type_wise_comparatives_pd.columns.values.tolist()
    print(columns_names)

    # create a new column - Success
    plant_type_wise_comparatives_pd['Variance'] = 0

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

    # variance formula implementation using index
    for index in plant_type_wise_comparatives_pd.index:
        present_quarter = plant_type_wise_comparatives_pd[columns_names[1]][index]
        previous_quarter = plant_type_wise_comparatives_pd[columns_names[2]][index]
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        plant_type_wise_comparatives_pd['Variance'][index] = variance

    # sorting
    # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
    grand_total = plant_type_wise_comparatives_pd[columns_names[1]].values[-1]
    plant_type_wise_comparatives_pd[columns_names[1]].values[-1] = 0
    plant_type_wise_comparatives_pd.sort_values(by=columns_names[1], axis=0, ascending=False, inplace=True)

    plant_type_wise_comparatives_pd[columns_names[1]].values[-1] = grand_total

    return plant_type_wise_comparatives_pd


def generate_domestic_and_import_wise(excel_file_pd, previous_quarter_final_file):

    # create a new column 'Purchase Type' with blank value
    excel_file_pd['Purchase Type'] = ''

    # Setting Type of purchase column values using currency key column on condition
    excel_file_pd.loc[excel_file_pd['Currency Key'] == "INR", 'Purchase Type'] = "Domestic"
    excel_file_pd.loc[excel_file_pd['Currency Key'] != "INR", 'Purchase Type'] = "Import"

    #  selecting only required columns
    excel_file_pd = excel_file_pd[['Purchase Type', 'GR Amt.in loc.cur.']]

    # create pivot table - sorting not required
    domestic_and_import_wise_pd = pd.pivot_table(excel_file_pd, index=["Purchase Type"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")

    # reset month index after pivot table creation for concatenation
    domestic_and_import_wise_pd = domestic_and_import_wise_pd.reset_index()

    # read previous quarters final working file
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Dom&Imp Wise Comparatives", usecols="A,C")

    # merging present and previous quarter purchase type wise data
    domestic_and_import_wise_comparatives_pd = pd.merge(domestic_and_import_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Purchase Type"])

    columns_list = domestic_and_import_wise_comparatives_pd.columns.values.tolist()
    print(columns_list)
    # create a new column - Success
    domestic_and_import_wise_comparatives_pd['Variance'] = 0

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None  # modifying only one df, so suppressing this warning as it is not affecting

    # variance formula implementation using index
    for index in domestic_and_import_wise_comparatives_pd.index:
        present_quarter = domestic_and_import_wise_comparatives_pd[columns_list[1]][index]
        previous_quarter = domestic_and_import_wise_comparatives_pd[columns_list[2]][index]
        variance = (present_quarter - previous_quarter) / previous_quarter
        domestic_and_import_wise_comparatives_pd['Variance'][index] = variance

    return domestic_and_import_wise_comparatives_pd


def find_financial_year_and_quarter(excel_file_pd):
    excel_file_pd['Year'] = pd.DatetimeIndex(excel_file_pd['GR Posting Date']).year
    year = 0  # initializing year
    financial_quarter = ''
    #  selecting only required columns
    excel_file_pd = excel_file_pd[["Month", "Year"]]

    months = excel_file_pd['Month'].unique().tolist()    # output example ['Jan', 'Feb', 'Mar']

    years = excel_file_pd['Year'].unique().tolist()   # output example [2022]

    if len(years) == 1:
        year = years[0]

    # find quarter
    if months == ['Jan', 'Feb', 'Mar']:
        quarter = 'Q4'
    elif months == ['Apr', 'May', 'Jun']:
        quarter = 'Q1'
    elif months == ['Jul', 'Aug', 'Sep']:
        quarter = 'Q2'
    elif months == ['Oct', 'Nov', 'Dec']:
        quarter = 'Q3'
    else:
        quarter = ''
    # determine financial year
    if quarter == 'Q1' or quarter == 'Q2' or quarter == 'Q3':
        next_year = year + 1
        financial_quarter = quarter + ' ' + 'FY' + ' ' + str(year % 2000) + "-" + str(next_year % 2000)
    elif quarter == 'Q4':
        previous_year = year - 1
        financial_quarter = quarter + ' ' + 'FY' + ' ' + str(previous_year % 2000) + "-" + str(year % 2000)
    else:
        pass

    return financial_quarter


def financial_quarter_title_update(file, quarter):
    workbook = openpyxl.load_workbook(file)
    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    worksheet_purchase_type.cell(1, 3).value = quarter      # row 1 column 3 = C1 cell value update

    worksheet_month_wise = workbook['Month Wise Comparatives']
    worksheet_month_wise.cell(1, 2).value = quarter         # B1 cell
    worksheet_month_wise.cell(1, 3).value = "Month"         # C1 cell - Its value is changed to Month.1 while processing, adjusting it back to "Month"

    worksheet_plant_type = workbook['Plant Wise Comparatives']
    worksheet_plant_type.cell(1, 2).value = quarter         # B1 cell

    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']
    worksheet_domestic_import_wise.cell(1, 2).value = quarter  # B1 cell

    workbook.save(file)


def number_formatting(file):
    workbook = openpyxl.load_workbook(file)
    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    worksheet_month_wise = workbook['Month Wise Comparatives']
    worksheet_plant_type = workbook['Plant Wise Comparatives']
    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']

    # Purchase type formatting
    for cell in worksheet_purchase_type['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_purchase_type['D']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_purchase_type['E']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # month wise formatting
    for cell in worksheet_month_wise['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_month_wise['D']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_month_wise['E']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # Plant type formatting
    for cell in worksheet_plant_type['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_plant_type['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_plant_type['D']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # Domestic and import type formatting
    for cell in worksheet_domestic_import_wise['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_domestic_import_wise['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_domestic_import_wise['D']:
        cell.number_format = '##.##%'

    workbook.save(file)


def apply_font_name_size_column_width(file):
    font_style = Font(name="Cambria", size=11)
    workbook = openpyxl.load_workbook(file)

    # purchase type
    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    worksheet_purchase_type.column_dimensions['A'].width = 20
    worksheet_purchase_type.column_dimensions['B'].width = 20
    worksheet_purchase_type.column_dimensions['C'].width = 20
    worksheet_purchase_type.column_dimensions['D'].width = 20
    worksheet_purchase_type.column_dimensions['E'].width = 15
    max_rows = worksheet_purchase_type.max_row
    max_cols = worksheet_purchase_type.max_column
    for i in range(1, max_rows+1):
        for j in range(1, max_cols+1):
            worksheet_purchase_type.cell(i, j).font = font_style

    # Month
    worksheet_month_wise = workbook['Month Wise Comparatives']
    worksheet_month_wise.column_dimensions['A'].width = 20
    worksheet_month_wise.column_dimensions['B'].width = 20
    worksheet_month_wise.column_dimensions['C'].width = 20
    worksheet_month_wise.column_dimensions['D'].width = 20
    worksheet_month_wise.column_dimensions['E'].width = 15
    max_rows = worksheet_month_wise.max_row
    max_cols = worksheet_month_wise.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_month_wise.cell(i, j).font = font_style

    # plant
    worksheet_plant_type = workbook['Plant Wise Comparatives']
    worksheet_plant_type.column_dimensions['A'].width = 20
    worksheet_plant_type.column_dimensions['B'].width = 20
    worksheet_plant_type.column_dimensions['C'].width = 20
    worksheet_plant_type.column_dimensions['D'].width = 15
    max_rows = worksheet_plant_type.max_row
    max_cols = worksheet_plant_type.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_plant_type.cell(i, j).font = font_style

    # domestic
    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']
    worksheet_domestic_import_wise.column_dimensions['A'].width = 20
    worksheet_domestic_import_wise.column_dimensions['B'].width = 20
    worksheet_domestic_import_wise.column_dimensions['C'].width = 20
    worksheet_domestic_import_wise.column_dimensions['D'].width = 15
    max_rows = worksheet_domestic_import_wise.max_row
    max_cols = worksheet_domestic_import_wise.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_domestic_import_wise.cell(i, j).font = font_style

    workbook.save(file)


def bold_and_color_fill(file):
    workbook = openpyxl.load_workbook(file)

    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    max_rows = worksheet_purchase_type.max_row
    max_cols = worksheet_purchase_type.max_column
    for i in range(1, max_cols+1):
        worksheet_purchase_type.cell(1, i).font = Font(bold=True)
        worksheet_purchase_type.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_purchase_type.cell(max_rows, i).font = Font(bold=True)

    worksheet_month_wise = workbook['Month Wise Comparatives']
    max_rows = worksheet_month_wise.max_row
    max_cols = worksheet_month_wise.max_column
    for i in range(1, max_cols+1):
        worksheet_month_wise.cell(1, i).font = Font(bold=True)
        worksheet_month_wise.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_month_wise.cell(max_rows, i).font = Font(bold=True)

    worksheet_plant_type = workbook['Plant Wise Comparatives']
    max_rows = worksheet_plant_type.max_row
    max_cols = worksheet_plant_type.max_column
    for i in range(1, max_cols+1):
        worksheet_plant_type.cell(1, i).font = Font(bold=True)
        worksheet_plant_type.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_plant_type.cell(max_rows, i).font = Font(bold=True)

    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']
    max_rows = worksheet_domestic_import_wise.max_row
    max_cols = worksheet_domestic_import_wise.max_column
    for i in range(1, max_cols+1):
        worksheet_domestic_import_wise.cell(1, i).font = Font(bold=True)
        worksheet_domestic_import_wise.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_domestic_import_wise.cell(max_rows, i).font = Font(bold=True)

    workbook.save(file)


def main():
    print("starting purchase type wise at " + datetime.now().strftime("%H:%M:%S"))
    purchase_type_wise_output = generate_purchase_type_wise(excel_file_as_pd, previous_quarter_file)
    print("Completed purchase type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting month wise at " + datetime.now().strftime("%H:%M:%S"))
    month_wise_output = generate_month_wise(excel_file_as_pd, previous_quarter_file)
    print("Completed month wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting plant type wise at " + datetime.now().strftime("%H:%M:%S"))
    plant_type_wise_output = generate_plant_type_wise(excel_file_as_pd, previous_quarter_file)
    print("Completed plant type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))
    domestic_and_import_wise_output = generate_domestic_and_import_wise(excel_file_as_pd, previous_quarter_file)
    print("Completed Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))

    # save all output dataframes in a single excel file
    with pd.ExcelWriter("Output.xlsx", engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}}, mode='w') as writer:
        purchase_type_wise_output.to_excel(writer, sheet_name="Purchase Type Wise Comparatives", index=False)
        month_wise_output.to_excel(writer, sheet_name="Month Wise Comparatives", index=False)
        plant_type_wise_output.to_excel(writer, sheet_name="Plant Wise Comparatives", index=False)
        domestic_and_import_wise_output.to_excel(writer, sheet_name="Dom&Imp Wise Comparatives", index=False)

    financial_quarter = find_financial_year_and_quarter(excel_file_as_pd)

    financial_quarter_title_update(file="Output.xlsx", quarter=financial_quarter)

    number_formatting(file="Output.xlsx")

    apply_font_name_size_column_width(file="Output.xlsx")

    bold_and_color_fill(file="Output.xlsx")

    print("Completed at " + datetime.now().strftime("%H:%M:%S"))


main()
