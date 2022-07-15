import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import pandas as pd
import numpy
import time

Excel_path = "purchase_registers_raw.xlsx"
font_name = "Cambria"
font_size = "11"
full_border = True
add_filter = True


def generate_purchase_type_wise(path):
    excel_file = path
    # read Excel file using pandas
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")
    print(excel_file_pd.head())

    # round entire Excel file - no use
#    excel_file_pd.round(decimals=2)

#   rounding amounts column to 2 decimals testing - no difference
#    print(excel_file_pd['GR Amt.in loc.cur.'].head())
#    excel_file_pd['GR Amt.in loc.cur.'].round(decimals=2)
#    print(excel_file_pd['GR Amt.in loc.cur.'].head())

    # create pivot table

    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"],
                                           values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True,
                                           margins_name="Grand Total")
#    print(purchase_type_wise_pd)

    # convert column 'GR Amt.in loc.cur.' values to integer format from exponential format
    purchase_type_wise_pd['GR Amt.in loc.cur.'] = purchase_type_wise_pd['GR Amt.in loc.cur.'].astype('int64')

    # In column "Valuation Class Text", replace blank cell with NaN value to eliminate error getting duplicate lines while merging with previous quarter
    # purchase_type_wise_pd = purchase_type_wise_pd.replace('', numpy.nan, regex=True)
    # purchase_type_wise_pd = purchase_type_wise_pd.mask(purchase_type_wise_pd == '')
    # purchase_type_wise_pd = purchase_type_wise_pd.replace(r'^\s*$', numpy.nan, regex=True)

    # print pivot table
    print(purchase_type_wise_pd)

    # read previous quarters final working file, previous quarter column and to int format from exponential format
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D") #.astype('int64')


    # replace Nan with
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan,'', regex=True)
    print(previous_quarter_final_file_pd)

    result = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])
    print(result)
    # Not-Working append previous quarter data to present output sheet
#    purchase_type_wise_pd = purchase_type_wise_pd.join(previous_quarter_final_file_pd)
#    print(purchase_type_wise_pd)


#   testing merge
#    purchase_type_wise_pd = purchase_type_wise_pd.merge(previous_quarter_final_file_pd, how="outer", left_on=True)
#    print(purchase_type_wise_pd)
    # testing concat method - Not Working
#    print(pd.concat([purchase_type_wise_pd, previous_quarter_final_file_pd], axis=1, ignore_index=True))



    # Save output to a new file
    purchase_type_wise_pd.to_excel("Output.xlsx", sheet_name="Purchase Type Wise Comparatives")

    return purchase_type_wise_pd


def generate_month_wise(formatted_excel):
    month_wise_sheet = 0
    return month_wise_sheet


def generate_import_and_domestic_wise(formatted_excel):
    import_and_domestic_wise = 0
    return import_and_domestic_wise


def generate_plant_wise(formatted_excel):
    plant_wise = 0
    return plant_wise


def create_comparative_output_sheet(purchase_type_wise_sheet, month_wise_sheet, import_and_domestic_wise, plant_wise):
    pass


def main():
    #    print("starting read_and_number_formatting_excel function at " + datetime.now().strftime("%H:%M:%S"))
    #    sheet_obj = read_and_number_formatting_excel(Excel_path)
    #    print("Execution of read_format_excel function is completed at " + datetime.now().strftime("%H:%M:%S"))
    print("starting purchase type wise at " + datetime.now().strftime("%H:%M:%S"))
    purchase_type_wise_sheet = generate_purchase_type_wise(Excel_path)
    print("Completed purchase type wise at " + datetime.now().strftime("%H:%M:%S"))


#   month_wise_sheet = generate_month_wise(formatted_excel)
#   import_and_domestic_wise = generate_import_and_domestic_wise(formatted_excel)
#    plant_wise = generate_plant_wise(formatted_excel)
#    create_comparative_output_sheet(purchase_type_wise_sheet, month_wise_sheet, import_and_domestic_wise, plant_wise)


main()


def extra_methods(sheet_obj, wb_obj):
    last_cell = sheet_obj.max_column + sheet_obj.max_row

    # formatting without iteration
    print("formatting without iteration")

    # Formatting font style to every cell

    sheet_obj.cell('A1:BH9448').font = Font(size=11, name="Cambria")
    wb_obj.save("copy of purchase registers.xlsx")
    print("formatting without iteration is complete and saved the file")


def read_and_number_formatting_excel(path):
    # To open the workbook, workbook object is created
    print("Copy of raw_file started at " + datetime.now().strftime("%H:%M:%S"))
    wb_obj = openpyxl.load_workbook(path)
    wb_obj.save("copy of purchase registers.xlsx")
    wb_obj.close()
    print("copy of raw_file completed at " + datetime.now().strftime("%H:%M:%S"))

    wb_obj = openpyxl.load_workbook("copy of purchase registers.xlsx")
    # Get workbook active sheet object from the active attribute
    sheet_obj = wb_obj.active
    print("loaded new workbook at " + datetime.now().strftime("%H:%M:%S"))

    print("Formatting of the file is skipped")
    return sheet_obj