import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import pandas as pd
import numpy
import time

Excel_path = "purchase_registers_raw.xlsx"
font_name = "Cambria"
font_size = "11"
full_border = True
add_filter = True


def generate_purchase_type_wise(path):
    excel_file = path

    # read Excel file using pandas
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")
    # print(excel_file_pd.head())   # reading check - success

    # create pivot table
    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")
    # print(purchase_type_wise_pd) # - Pivot check -Success

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D")
    # print(previous_quarter_final_file_pd) # Success - prints with Nan

    # replace Nan with blank
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)
    # print(previous_quarter_final_file_pd)   # Success - replaces Nan with blank ''

    # merging present and previous quarter purchase type wise data
    purchase_type_wise_comparatives_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])

    # replacing all Nan's with zeros
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    # Converting numbers from exponential to number
    purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."] = purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."].astype('int64')
    purchase_type_wise_comparatives_pd["Q3 FY 21-22"] = purchase_type_wise_comparatives_pd["Q3 FY 21-22"].astype('int64')

    # apply comma format to numbers
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.style.format('{:,}')

    print(purchase_type_wise_comparatives_pd)

    # Save purchase type wise comparatives to output file
    purchase_type_wise_comparatives_pd.to_excel("Output.xlsx", sheet_name="Purchase Type Wise Comparatives")




def generate_month_wise(formatted_excel):
    month_wise_sheet = 0
    return month_wise_sheet


def generate_import_and_domestic_wise(formatted_excel):
    import_and_domestic_wise = 0
    return import_and_domestic_wise


def generate_plant_wise(formatted_excel):
    plant_wise = 0
    return plant_wise


def create_comparative_output_sheet(purchase_type_wise_sheet, month_wise_sheet, import_and_domestic_wise, plant_wise):
    pass


def main():
    #    print("starting read_and_number_formatting_excel function at " + datetime.now().strftime("%H:%M:%S"))
    #    sheet_obj = read_and_number_formatting_excel(Excel_path)
    #    print("Execution of read_format_excel function is completed at " + datetime.now().strftime("%H:%M:%S"))
    print("starting purchase type wise at " + datetime.now().strftime("%H:%M:%S"))
    generate_purchase_type_wise(Excel_path)
    print("Completed purchase type wise at " + datetime.now().strftime("%H:%M:%S"))


#   month_wise_sheet = generate_month_wise(formatted_excel)
#   import_and_domestic_wise = generate_import_and_domestic_wise(formatted_excel)
#    plant_wise = generate_plant_wise(formatted_excel)
#    create_comparative_output_sheet(purchase_type_wise_sheet, month_wise_sheet, import_and_domestic_wise, plant_wise)


main()


def extra_methods(sheet_obj, wb_obj):
    last_cell = sheet_obj.max_column + sheet_obj.max_row

    # formatting without iteration
    print("formatting without iteration")

    # Formatting font style to every cell

    sheet_obj.cell('A1:BH9448').font = Font(size=11, name="Cambria")
    wb_obj.save("copy of purchase registers.xlsx")
    print("formatting without iteration is complete and saved the file")
    # ---------------------------------------------------------------------------
    # round entire Excel file - no use
    #    excel_file_pd.round(decimals=2)
    #   rounding amounts column to 2 decimals testing - no difference
    #    print(excel_file_pd['GR Amt.in loc.cur.'].head())
    #    excel_file_pd['GR Amt.in loc.cur.'].round(decimals=2)
    #    print(excel_file_pd['GR Amt.in loc.cur.'].head())
    # -----------------------------------------------------------
    # In column "Valuation Class Text", replace blank cell with NaN value to eliminate error getting duplicate lines while merging with previous quarter
    # purchase_type_wise_pd = purchase_type_wise_pd.replace('', numpy.nan, regex=True)
    # purchase_type_wise_pd = purchase_type_wise_pd.mask(purchase_type_wise_pd == '')
    # purchase_type_wise_pd = purchase_type_wise_pd.replace(r'^\s*$', numpy.nan, regex=True)

    # -----------------------------------------------------------
    # convert column 'GR Amt.in loc.cur.' values to integer format from exponential format
    # purchase_type_wise_pd['GR Amt.in loc.cur.'] = purchase_type_wise_pd['GR Amt.in loc.cur.'].astype('int64')
    # Getting Previous Quarter values as exponential - converting to int
    # previous_quarter_final_file_pd['Q3 FY 21-22'] = previous_quarter_final_file_pd['Q3 FY 21-22'].astype('int64')
    # print(previous_quarter_final_file_pd)   # Success - Converting to it

    # ------------------------------------------------------------



def read_and_number_formatting_excel(path):
    # To open the workbook, workbook object is created
    print("Copy of raw_file started at " + datetime.now().strftime("%H:%M:%S"))
    wb_obj = openpyxl.load_workbook(path)
    wb_obj.save("copy of purchase registers.xlsx")
    wb_obj.close()
    print("copy of raw_file completed at " + datetime.now().strftime("%H:%M:%S"))

    wb_obj = openpyxl.load_workbook("copy of purchase registers.xlsx")
    # Get workbook active sheet object from the active attribute
    sheet_obj = wb_obj.active
    print("loaded new workbook at " + datetime.now().strftime("%H:%M:%S"))

    print("Formatting of the file is skipped")
    return sheet_obj