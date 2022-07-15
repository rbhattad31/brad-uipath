import openpyxl
from openpyxl.styles import Font, numbers
from openpyxl import workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import numpy
import time

pd.options.display.float_format = '{:,.2f}'.format  # excellent - comma, expo to int format and all at once

Excel_path = "purchase_registers_raw.xlsx"
font_name = "Cambria"
font_size = "11"
full_border = True


def generate_purchase_type_wise(path):
    excel_file = path

    # read Excel file using pandas
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")
    # print(excel_file_pd.head())   # reading check - success

    # create pivot table
    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=True)
    # print(purchase_type_wise_pd)  # - Pivot check -Success

    # reset indices created after pivot table creation for merging
    purchase_type_wise_pd = purchase_type_wise_pd.reset_index()
    print(purchase_type_wise_pd)  # need to test this line - and adding index 0 1 2 3 ..

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D")
    # print(previous_quarter_final_file_pd)  # Success - prints with Nan

    # replace Nan with blank
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)
    # print(previous_quarter_final_file_pd)   # Success - replaces Nan with blank ''

    # merging present and previous quarter purchase type wise data
    purchase_type_wise_comparatives_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])

    # replacing all Nan's with zeros in Present and previous values columns
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    # Converting numbers from exponential to number - delete after testing
    # purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."] = purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."].astype('int64')
    # purchase_type_wise_comparatives_pd["Q3 FY 21-22"] = purchase_type_wise_comparatives_pd["Q3 FY 21-22"].astype('int64')

    # create a new column - Success
    purchase_type_wise_comparatives_pd['variance'] = 0
    # print(purchase_type_wise_comparatives_pd)  # variance column with 0's

    # ignoring index numbers and save to output excel
    purchase_type_wise_comparatives_pd.to_excel('Purchase_Output.xlsx', sheet_name="Purchase Type Wise Comparatives", index=False)

    # openpyxl starts here
    workbook_object = openpyxl.load_workbook("Purchase_Output.xlsx")
    purchase_type_wise_comparatives_worksheet = workbook_object["Purchase Type Wise Comparatives"]
    max_rows = purchase_type_wise_comparatives_worksheet.max_row

    for i in range(2, max_rows+1):
        present_quarter = purchase_type_wise_comparatives_worksheet.cell(row=i, column=3).value
        previous_quarter = purchase_type_wise_comparatives_worksheet.cell(row=i, column=4).value
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        purchase_type_wise_comparatives_worksheet.cell(row=i, column=5).value = variance
        # rounding and limit decimals to two places in percentage format
        purchase_type_wise_comparatives_worksheet.cell(row=i, column=5).number_format = '0.00%'

    # number format comma separated - Pending
    # style formatting - Pending
    # Rows with zero values for 2 Quarters - Pending
    # sorting from high to low - Pending

    workbook_object.save("Purchase_Output.xlsx")
    print("Purchase Type Wise Comparatives is created in output.xlsx")


def generate_plant_type_wise(path):
    excel_file = path
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")

    # create pivot table
    plant_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Plant"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=True)
    #  print(plant_type_wise_pd)  # - Pivot check -Success

    # reset index created after pivot table creation for merging
    plant_type_wise_pd = plant_type_wise_pd.reset_index()
    print(plant_type_wise_pd)  # need to test this line - and adding index 0 1 2 3 ..

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Plant Wise Comparatives", usecols="A,C")
    #  print(previous_quarter_final_file_pd)

    # merging present and previous quarter purchase type wise data
    plant_type_wise_comparatives_pd = pd.merge(plant_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Plant"])
    #  print(plant_type_wise_comparatives_pd)

    # replacing all Nan's with zeros in Present and previous values columns - not necessary but if a new plant is added
    # eliminates NaN error
    plant_type_wise_comparatives_pd = plant_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    # Converting numbers from exponential to number - delete after testing
    # plant_type_wise_comparatives_pd["GR Amt.in loc.cur."] = plant_type_wise_comparatives_pd["GR Amt.in loc.cur."].astype('int64')
    # plant_type_wise_comparatives_pd["Q3 FY 21-22"] = plant_type_wise_comparatives_pd["Q3 FY 21-22"].astype('int64')

    # create a new column - Success
    plant_type_wise_comparatives_pd['variance'] = 0
    #  print(plant_type_wise_comparatives_pd)  # variance column with 0's

    # ignoring index numbers and save to output excel
    # with pd.ExcelWriter("Output.xlsx") as writer:
    #     writer.book = openpyxl.load_workbook("Output.xlsx")
    #     plant_type_wise_comparatives_pd.to_excel(writer, sheet_name="Plant Wise Comparatives", index=False)

    plant_type_wise_comparatives_pd.to_excel('Plant_Output.xlsx', sheet_name="Plant Wise Comparatives", index=False)

    # openpyxl starts here
    workbook_object = openpyxl.load_workbook("Plant_Output.xlsx")
    plant_type_wise_comparatives_worksheet = workbook_object["Plant Wise Comparatives"]
    max_rows = plant_type_wise_comparatives_worksheet.max_row

    for i in range(2, max_rows+1):
        present_quarter = plant_type_wise_comparatives_worksheet.cell(row=i, column=2).value
        previous_quarter = plant_type_wise_comparatives_worksheet.cell(row=i, column=3).value
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        plant_type_wise_comparatives_worksheet.cell(row=i, column=4).value = variance
        # rounding and limit decimals to two places in percentage format
        plant_type_wise_comparatives_worksheet.cell(row=i, column=4).number_format = '0.00%'

    # number format comma separated - Pending
    # style formatting - Pending
    # Rows with zero values for 2 Quarters - Pending
    # sorting from high to low - Pending

    workbook_object.save("Plant_Output.xlsx")
    print("Plant Type Wise Comparatives is created in output.xlsx")


def generate_month_wise(path):
    excel_file = path
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")

    #  Create Month column
    excel_file_pd['Month'] = pd.DatetimeIndex(excel_file_pd['GR Posting Date']).month_name().str[:3]
    #  print(excel_file_pd.head())   # Having Jan, Feb , Mar  --- 3 letter format

    #  selecting only required columns
    excel_file_pd = excel_file_pd[["GR Posting Date", "GR Amt.in loc.cur.", "Month"]]

    # create pivot table
    month_wise_pd = pd.pivot_table(excel_file_pd, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=False)
    # print(month_wise_pd)  # - Pivot check -Success

    # reset month index after pivot table creation for concatenation
    month_wise_pd = month_wise_pd.reset_index()
    print(month_wise_pd)  # success - and adding index 0 1 2 3 ..

    # read from previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Month Wise Comparatives", usecols="C,D")
    print(previous_quarter_final_file_pd)

    # concatenation
    month_wise_comparatives_pd = pd.concat([month_wise_pd, previous_quarter_final_file_pd], axis=1)
    print(month_wise_comparatives_pd)

    # create a new column - Success
    month_wise_comparatives_pd['variance'] = 0
    print(month_wise_comparatives_pd)  # variance column with 0's

    # ignoring index numbers and save to output excel
    month_wise_comparatives_pd.to_excel('Month_Output.xlsx', sheet_name="Month Wise Comparatives", index=False)

    # variance formula implementation using openpyxl starts here
    workbook_object = openpyxl.load_workbook("Month_Output.xlsx")
    month_wise_comparatives_worksheet = workbook_object["Month Wise Comparatives"]
    month_wise_comparatives_worksheet.number_format = '{:,.2f}'.format

    max_rows = month_wise_comparatives_worksheet.max_row

    for i in range(2, max_rows+1):
        present_quarter = month_wise_comparatives_worksheet.cell(row=i, column=2).value
        previous_quarter = month_wise_comparatives_worksheet.cell(row=i, column=4).value
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        month_wise_comparatives_worksheet.cell(row=i, column=5).value = variance
        # rounding and limit decimals to two places in percentage format
        month_wise_comparatives_worksheet.cell(row=i, column=5).number_format = '0.00%'

    # number format comma separated - Pending
    # style formatting - Pending
    # Rows with zero values for 2 Quarters - Pending
    # sorting from high to low - Pending

    workbook_object.save("Month_Output.xlsx")
    print("Month Wise Comparatives is created in Month_Output.xlsx")


def generate_import_and_domestic_wise(path):
    excel_file = path

    # read Excel file using pandas
    excel_file_pd = pd.read_excel(excel_file, sheet_name="Sheet1")
    # print(excel_file_pd.head())   # reading check - success

    # create pivot table
    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")
    # print(purchase_type_wise_pd)  # - Pivot check -Success

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D")
    # print(previous_quarter_final_file_pd)  # Success - prints with Nan

    # replace Nan with blank
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)
    # print(previous_quarter_final_file_pd)   # Success - replaces Nan with blank ''

    # merging present and previous quarter purchase type wise data
    purchase_type_wise_comparatives_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])

    # replacing all Nan's with zeros in Present and previous values columns
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    # Converting numbers from exponential to number
    purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."] = purchase_type_wise_comparatives_pd["GR Amt.in loc.cur."].astype('int64')
    purchase_type_wise_comparatives_pd["Q3 FY 21-22"] = purchase_type_wise_comparatives_pd["Q3 FY 21-22"].astype('int64')

    # create a new column - Success
    purchase_type_wise_comparatives_pd['variance'] = 0
    # print(purchase_type_wise_comparatives_pd)  # variance column with 0's

    # ignoring index numbers and save to output excel
    purchase_type_wise_comparatives_pd.to_excel('Purchase_Output.xlsx', sheet_name="Purchase Type Wise Comparatives", index=False)

    # openpyxl starts here
    workbook_object = openpyxl.load_workbook("Purchase_Output.xlsx")
    purchase_type_wise_comparatives_worksheet = workbook_object["Purchase Type Wise Comparatives"]
    max_rows = purchase_type_wise_comparatives_worksheet.max_row

    for i in range(2, max_rows+1):
        present_quarter = purchase_type_wise_comparatives_worksheet.cell(row=i, column=3).value
        previous_quarter = purchase_type_wise_comparatives_worksheet.cell(row=i, column=4).value
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        purchase_type_wise_comparatives_worksheet.cell(row=i, column=5).value = variance
        # rounding and limit decimals to two places in percentage format
        purchase_type_wise_comparatives_worksheet.cell(row=i, column=5).number_format = '0.00%'

    # number format comma separated - Pending
    # style formatting - Pending
    # Rows with zero values for 2 Quarters - Pending
    # sorting from high to low - Pending

    workbook_object.save("Purchase_Output.xlsx")
    print("Purchase Type Wise Comparatives is created in output.xlsx")


def main():

    print("starting purchase type wise at " + datetime.now().strftime("%H:%M:%S"))
#    generate_purchase_type_wise(Excel_path)
    print("Completed purchase type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting month wise at " + datetime.now().strftime("%H:%M:%S"))
    generate_month_wise(Excel_path)
    print("Completed month wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting plant type wise at " + datetime.now().strftime("%H:%M:%S"))
#    generate_plant_type_wise(Excel_path)
    print("Completed plant type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))
    #  generate_month_wise(Excel_path)
    print("Completed Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))


main()




def extra_methods(sheet_obj, wb_obj):
    last_cell = sheet_obj.max_column + sheet_obj.max_row

    # formatting without iteration
    print("formatting without iteration")

    # Formatting font style to every cell

    sheet_obj.cell('A1:BH9448').font = Font(size=11, name="Cambria")
    wb_obj.save("copy of purchase registers.xlsx")
    print("formatting without iteration is complete and saved the file")
    # ---------------------------------------------------------------------------
    # round entire Excel file - no use
    #    excel_file_pd.round(decimals=2)
    #   rounding amounts column to 2 decimals testing - no difference
    #    print(excel_file_pd['GR Amt.in loc.cur.'].head())
    #    excel_file_pd['GR Amt.in loc.cur.'].round(decimals=2)
    #    print(excel_file_pd['GR Amt.in loc.cur.'].head())
    # -----------------------------------------------------------
    # In column "Valuation Class Text", replace blank cell with NaN value to eliminate error getting duplicate lines while merging with previous quarter
    # purchase_type_wise_pd = purchase_type_wise_pd.replace('', numpy.nan, regex=True)
    # purchase_type_wise_pd = purchase_type_wise_pd.mask(purchase_type_wise_pd == '')
    # purchase_type_wise_pd = purchase_type_wise_pd.replace(r'^\s*$', numpy.nan, regex=True)

    # -----------------------------------------------------------
    # convert column 'GR Amt.in loc.cur.' values to integer format from exponential format
    # purchase_type_wise_pd['GR Amt.in loc.cur.'] = purchase_type_wise_pd['GR Amt.in loc.cur.'].astype('int64')
    # Getting Previous Quarter values as exponential - converting to int
    # previous_quarter_final_file_pd['Q3 FY 21-22'] = previous_quarter_final_file_pd['Q3 FY 21-22'].astype('int64')
    # print(previous_quarter_final_file_pd)   # Success - Converting to it

    # ------------------------------------------------------------



def read_and_number_formatting_excel(path):
    # To open the workbook, workbook object is created
    print("Copy of raw_file started at " + datetime.now().strftime("%H:%M:%S"))
    wb_obj = openpyxl.load_workbook(path)
    wb_obj.save("copy of purchase registers.xlsx")
    wb_obj.close()
    print("copy of raw_file completed at " + datetime.now().strftime("%H:%M:%S"))

    wb_obj = openpyxl.load_workbook("copy of purchase registers.xlsx")
    # Get workbook active sheet object from the active attribute
    sheet_obj = wb_obj.active
    print("loaded new workbook at " + datetime.now().strftime("%H:%M:%S"))

    # print("starting read_and_number_formatting_excel function at " + datetime.now().strftime("%H:%M:%S"))
    # sheet_obj = read_and_number_formatting_excel(Excel_path)
    # print("Execution of read_format_excel function is completed at " + datetime.now().strftime("%H:%M:%S"))


    print("Formatting of the file is skipped")
    return sheet_obj