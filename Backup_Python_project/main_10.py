import pandas as pd
from datetime import datetime
import numpy
import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from xlsxwriter import Workbook

pd.options.display.float_format = '{:,.2f}'.format  # comma, expo to int format and all at once but failed to write them to excel

excel_path = "purchase_registers_raw.xlsx"

excel_file_as_pd = pd.read_excel(excel_path, sheet_name="Sheet1")


def generate_purchase_type_wise(excel_file_pd):
    # create pivot table
    purchase_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Valuation Class", "Valuation Class Text"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")
    print(purchase_type_wise_pd)  # - Pivot check -Success

    # reset "indices created during pivot table creation" - for merging
    purchase_type_wise_pd = purchase_type_wise_pd.reset_index()
    print(purchase_type_wise_pd)  # need to test this line - and adding index 0 1 2 3 etc

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Purchase Type Wise Comparatives", usecols="A,B,D")
    print(previous_quarter_final_file_pd)  # Success - prints with Nan

    # replace Nan with blank
    previous_quarter_final_file_pd = previous_quarter_final_file_pd.replace(numpy.nan, '', regex=True)
    print(previous_quarter_final_file_pd)   # Success - replaces Nan with blank ''

    # merging present and previous quarter purchase type wise data
    purchase_type_wise_comparatives_pd = pd.merge(purchase_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Valuation Class", "Valuation Class Text"])
    print(purchase_type_wise_comparatives_pd)

    # replacing all Nan's with zeros in Present and previous Quarter's values columns
    purchase_type_wise_comparatives_pd = purchase_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)
    print(purchase_type_wise_comparatives_pd)

    # create a new column - Success
    purchase_type_wise_comparatives_pd['variance'] = 0
    print(purchase_type_wise_comparatives_pd)  # variance column with 0's

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None

    # variance formula implementation using index
    for index in purchase_type_wise_comparatives_pd.index:
        present_quarter = purchase_type_wise_comparatives_pd['GR Amt.in loc.cur.'][index]
        previous_quarter = purchase_type_wise_comparatives_pd['Q3 FY 21-22'][index]
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        purchase_type_wise_comparatives_pd['variance'][index] = variance

    print(purchase_type_wise_comparatives_pd)

    # sorting
    # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
    grand_total = purchase_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1]
    print(grand_total)
    purchase_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1] = 0
    print(purchase_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1])
    purchase_type_wise_comparatives_pd.sort_values(by="GR Amt.in loc.cur.", axis=0, ascending=False, inplace=True)
    print(purchase_type_wise_comparatives_pd)

    purchase_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1] = grand_total
    print(purchase_type_wise_comparatives_pd)

    return purchase_type_wise_comparatives_pd


def generate_plant_type_wise(excel_file_pd):

    # create pivot table
    plant_type_wise_pd = pd.pivot_table(excel_file_pd, index=["Plant"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=True)
    print(plant_type_wise_pd)  # - Pivot check -Success

    # reset index created after pivot table creation for merging
    plant_type_wise_pd = plant_type_wise_pd.reset_index()
    print(plant_type_wise_pd)  # need to test this line - and adding index 0 1 2 3 etc

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Plant Wise Comparatives", usecols="A,C")
    print(previous_quarter_final_file_pd)

    # merging present and previous quarter purchase type wise data
    plant_type_wise_comparatives_pd = pd.merge(plant_type_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Plant"])
    print(plant_type_wise_comparatives_pd)

    # replacing all Nan's with zeros in Present and previous values columns - not necessary now but if a new plant is added
    # Nan's will be formed in previous quarter columns, eliminates NaN error
    plant_type_wise_comparatives_pd = plant_type_wise_comparatives_pd.replace(numpy.nan, 0, regex=True)

    # create a new column - Success
    plant_type_wise_comparatives_pd['variance'] = 0
    print(plant_type_wise_comparatives_pd)  # variance column with 0's

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None

    # variance formula implementation using index
    for index in plant_type_wise_comparatives_pd.index:
        present_quarter = plant_type_wise_comparatives_pd['GR Amt.in loc.cur.'][index]
        previous_quarter = plant_type_wise_comparatives_pd['Q3 FY 21-22'][index]
        if previous_quarter == 0:
            variance = 1
        else:
            variance = (present_quarter - previous_quarter) / previous_quarter
        plant_type_wise_comparatives_pd['variance'][index] = variance

    print(plant_type_wise_comparatives_pd)

    # sorting
    # copy present quarter Amount column Grand total, set it as zero, sort the data frame and reassign the value.
    grand_total = plant_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1]
    print(grand_total)
    plant_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1] = 0
    print(plant_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1])
    plant_type_wise_comparatives_pd.sort_values(by="GR Amt.in loc.cur.", axis=0, ascending=False, inplace=True)
    print(plant_type_wise_comparatives_pd)

    plant_type_wise_comparatives_pd['GR Amt.in loc.cur.'].values[-1] = grand_total
    print(plant_type_wise_comparatives_pd)

    return plant_type_wise_comparatives_pd


def generate_month_wise(excel_file_pd):

    #  Create Month column
    excel_file_pd['Month'] = pd.DatetimeIndex(excel_file_pd['GR Posting Date']).month_name().str[:3]

    #  selecting only required columns
    excel_file_pd = excel_file_pd[["GR Posting Date", "GR Amt.in loc.cur.", "Month"]]
    print(excel_file_pd)

    # create pivot table, sort = False to not sort Month column as per alphabetical order - mandatory
    month_wise_pd = pd.pivot_table(excel_file_pd, index=["Month"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total", sort=False)
    print(month_wise_pd)  # - Pivot check -Success

    # reset month index after pivot table creation for concatenation
    month_wise_pd = month_wise_pd.reset_index()
    print(month_wise_pd)  # success - and adding index 0 1 2 3 etc

    # read from previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Month Wise Comparatives", usecols="C,D")
    print(previous_quarter_final_file_pd)

    # concatenation instead of merge as there are no common Columns to merge.
    month_wise_comparatives_pd = pd.concat([month_wise_pd, previous_quarter_final_file_pd], axis=1)
    print(month_wise_comparatives_pd)

    # create a new column - Success
    month_wise_comparatives_pd['variance'] = 0
    print(month_wise_comparatives_pd)  # variance column with 0's

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None

    # variance formula implementation using index
    for index in month_wise_comparatives_pd.index:
        present_quarter = month_wise_comparatives_pd['GR Amt.in loc.cur.'][index]
        previous_quarter = month_wise_comparatives_pd['Q3 FY 21-22'][index]
        variance = (present_quarter - previous_quarter) / previous_quarter
        month_wise_comparatives_pd['variance'][index] = variance

    print(month_wise_comparatives_pd)

    return month_wise_comparatives_pd


def generate_domestic_and_import_wise(excel_file_pd):

    # create a new column 'Purchase Type' with blank value
    excel_file_pd['Purchase Type'] = ''

    # Setting Type of purchase column values using currency key column on condition
    excel_file_pd.loc[excel_file_pd['Currency Key'] == "INR", 'Purchase Type'] = "Domestic"
    excel_file_pd.loc[excel_file_pd['Currency Key'] != "INR", 'Purchase Type'] = "Import"

    #  selecting only required columns
    excel_file_pd = excel_file_pd[['Purchase Type', 'GR Amt.in loc.cur.']]
    print(excel_file_pd)

    # create pivot table - sorting not required
    domestic_and_import_wise_pd = pd.pivot_table(excel_file_pd, index=["Purchase Type"], values="GR Amt.in loc.cur.", aggfunc=numpy.sum, margins=True, margins_name="Grand Total")
    print(domestic_and_import_wise_pd)  # - Pivot check -Success

    # reset month index after pivot table creation for concatenation
    domestic_and_import_wise_pd = domestic_and_import_wise_pd.reset_index()
    print(domestic_and_import_wise_pd)  # success - and adding index 0 1 2 3 etc

    # read previous quarters final working file
    previous_quarter_final_file = "Previous_Quarter_Final_File.xlsx"
    previous_quarter_final_file_pd = pd.read_excel(previous_quarter_final_file, sheet_name="Dom&Imp Wise Comparatives", usecols="A,C")
    print(previous_quarter_final_file_pd)  # Success

    # merging present and previous quarter purchase type wise data
    domestic_and_import_wise_comparatives_pd = pd.merge(domestic_and_import_wise_pd, previous_quarter_final_file_pd, how="outer", on=["Purchase Type"])
    print(domestic_and_import_wise_comparatives_pd)

    # create a new column - Success
    domestic_and_import_wise_comparatives_pd['variance'] = 0
    print(domestic_and_import_wise_comparatives_pd)  # variance column with 0's

    # To Remove SettingWithCopyWarning error
    pd.options.mode.chained_assignment = None

    # variance formula implementation using index
    for index in domestic_and_import_wise_comparatives_pd.index:
        present_quarter = domestic_and_import_wise_comparatives_pd['GR Amt.in loc.cur.'][index]
        previous_quarter = domestic_and_import_wise_comparatives_pd['Q3 FY 21-22'][index]
        variance = (present_quarter - previous_quarter) / previous_quarter
        domestic_and_import_wise_comparatives_pd['variance'][index] = variance

    print(domestic_and_import_wise_comparatives_pd)

    return domestic_and_import_wise_comparatives_pd


def number_formatting(file):
    workbook = openpyxl.load_workbook(file)
    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    worksheet_month_wise = workbook['Month Wise Comparatives']
    worksheet_plant_type = workbook['Plant Wise Comparatives']
    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']

    # Purchase type formatting
    for cell in worksheet_purchase_type['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_purchase_type['D']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_purchase_type['E']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # month wise formatting
    for cell in worksheet_month_wise['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_month_wise['D']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_month_wise['E']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # Plant type formatting
    for cell in worksheet_plant_type['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_plant_type['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_plant_type['D']:
        cell.number_format = '##.##%'
    # -------------------------------------------------

    # Domestic and import type formatting
    for cell in worksheet_domestic_import_wise['B']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_domestic_import_wise['C']:
        if cell == 0.00:
            continue
        else:
            cell.number_format = '#,###'

    for cell in worksheet_domestic_import_wise['D']:
        cell.number_format = '##.##%'

    workbook.save(file)


def apply_font_and_size(file):
    font_style = Font(name="Cambria", size=11)
    workbook = openpyxl.load_workbook(file)

    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    max_rows = worksheet_purchase_type.max_row
    max_cols = worksheet_purchase_type.max_column
    for i in range(1, max_rows+1):
        for j in range(1, max_cols+1):
            worksheet_purchase_type.cell(i, j).font = font_style

    worksheet_month_wise = workbook['Month Wise Comparatives']
    max_rows = worksheet_month_wise.max_row
    max_cols = worksheet_month_wise.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_month_wise.cell(i, j).font = font_style

    worksheet_plant_type = workbook['Plant Wise Comparatives']
    max_rows = worksheet_plant_type.max_row
    max_cols = worksheet_plant_type.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_plant_type.cell(i, j).font = font_style

    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']
    max_rows = worksheet_domestic_import_wise.max_row
    max_cols = worksheet_domestic_import_wise.max_column
    for i in range(1, max_rows + 1):
        for j in range(1, max_cols + 1):
            worksheet_domestic_import_wise.cell(i, j).font = font_style

    workbook.save(file)


def bold_and_color_fill(file):
    workbook = openpyxl.load_workbook(file)

    worksheet_purchase_type = workbook['Purchase Type Wise Comparatives']
    max_rows = worksheet_purchase_type.max_row
    max_cols = worksheet_purchase_type.max_column
    for i in range(1, max_cols+1):
        worksheet_purchase_type.cell(1, i).font = Font(bold=True)
        worksheet_purchase_type.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_purchase_type.cell(max_rows, i).font = Font(bold=True)

    worksheet_month_wise = workbook['Month Wise Comparatives']
    max_rows = worksheet_month_wise.max_row
    max_cols = worksheet_month_wise.max_column
    for i in range(1, max_cols+1):
        worksheet_month_wise.cell(1, i).font = Font(bold=True)
        worksheet_month_wise.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_month_wise.cell(max_rows, i).font = Font(bold=True)

    worksheet_plant_type = workbook['Plant Wise Comparatives']
    max_rows = worksheet_plant_type.max_row
    max_cols = worksheet_plant_type.max_column
    for i in range(1, max_cols+1):
        worksheet_plant_type.cell(1, i).font = Font(bold=True)
        worksheet_plant_type.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_plant_type.cell(max_rows, i).font = Font(bold=True)

    worksheet_domestic_import_wise = workbook['Dom&Imp Wise Comparatives']
    max_rows = worksheet_domestic_import_wise.max_row
    max_cols = worksheet_domestic_import_wise.max_column
    for i in range(1, max_cols+1):
        worksheet_domestic_import_wise.cell(1, i).font = Font(bold=True)
        worksheet_domestic_import_wise.cell(1, i).fill = PatternFill(start_color='00a2ed', fill_type='solid')
        worksheet_domestic_import_wise.cell(max_rows, i).font = Font(bold=True)

    workbook.save(file)


def main():
    print("starting purchase type wise at " + datetime.now().strftime("%H:%M:%S"))
    purchase_type_wise_output = generate_purchase_type_wise(excel_file_as_pd)
    print("Completed purchase type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting month wise at " + datetime.now().strftime("%H:%M:%S"))
    month_wise_output = generate_month_wise(excel_file_as_pd)
    print("Completed month wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting plant type wise at " + datetime.now().strftime("%H:%M:%S"))
    plant_type_wise_output = generate_plant_type_wise(excel_file_as_pd)
    print("Completed plant type wise at " + datetime.now().strftime("%H:%M:%S"))

    print("starting Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))
    domestic_and_import_wise_output = generate_domestic_and_import_wise(excel_file_as_pd)
    print("Completed Domestic and import type wise at " + datetime.now().strftime("%H:%M:%S"))

    # save all output dataframes in a single excel file
    with pd.ExcelWriter("Output.xlsx", engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}}, mode='w') as writer:
        purchase_type_wise_output.to_excel(writer, sheet_name="Purchase Type Wise Comparatives", index=False)
        month_wise_output.to_excel(writer, sheet_name="Month Wise Comparatives", index=False)
        plant_type_wise_output.to_excel(writer, sheet_name="Plant Wise Comparatives", index=False)
        domestic_and_import_wise_output.to_excel(writer, sheet_name="Dom&Imp Wise Comparatives", index=False)

    number_formatting(file="Output.xlsx")

    apply_font_and_size(file="Output.xlsx")

    bold_and_color_fill(file="Output.xlsx")




main()
